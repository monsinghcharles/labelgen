from django.shortcuts import render

from django.shortcuts import render
import openpyxl
from django.http.response import HttpResponse
import qrcode
from PIL import Image, ImageDraw, ImageFont
import PIL.Image
import textwrap
import pandas
from PyPDF2 import PdfFileMerger, PdfFileReader
import tempfile
from tkinter import *
import os
import mimetypes
from PIL import Image, ImageChops
from PIL import Image, ImageOps



def indix(request):

    temp_dir = tempfile.mkdtemp()
    wlf_path = 'media/labelsrc/wlf.png'
    hbd_path = 'media/labelsrc/birthday.jpg'
    iro_path = 'media/labelsrc/iro.jpg'
    anni_path = 'media/labelsrc/anni.jpg'
    arial_path = 'media/labelsrc/arial-unicode-ms.ttf'
    bebas_path = 'media/labelsrc/BebasNeue-Regular.ttf'
    white_path = 'media/labelsrc/white.jpg'
    y = 0

    def MakeQR(nameop):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=17,
            border=1,
        )
        # nameop = "thaagam.org/qr/{}".format(nameop)
        # nameop = nameop
        # nameop = nameop.replace(" ","-")
        if len(nameop) < 4:
            nameop = nameop + "---"
        qr.add_data(nameop)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        width, height = img.size
        img = img.resize((388, 388))
        return img
    def trim(im):
        bg = Image.new(im.mode, im.size, im.getpixel((0, 0)))
        diff = ImageChops.difference(im, bg)
        diff = ImageChops.add(diff, diff, 2.0, -100)
        bbox = diff.getbbox()
        if bbox:
            return im.crop(bbox)
    def OShortName(name):
        import re
        name = str(name)
        hbdre = re.compile(re.escape("Happy Birthday"), re.IGNORECASE)
        name = hbdre.sub("<HBD>", name)

        haare = re.compile(re.escape("Happy anniversary"), re.IGNORECASE)
        name = haare.sub("<HAA>", name)

        irore = re.compile(re.escape("in remembrance of"), re.IGNORECASE)
        name = irore.sub("<IRO>", name)

        wlfre = re.compile(re.escape("With love from"), re.IGNORECASE)
        name = wlfre.sub(" ", name)

        return name
    def resize_with_padding(img, expected_size):
        img.thumbnail((expected_size[0], expected_size[1]))
        # print(img.size)
        delta_width = expected_size[0] - img.size[0]
        delta_height = expected_size[1] - img.size[1]
        pad_width = delta_width // 2
        pad_height = delta_height // 2
        padding = (pad_width, pad_height, delta_width - pad_width, delta_height - pad_height)
        return ImageOps.expand(img, padding, fill=(255, 255, 255))
    def Occationsettings(name):
        LH = len(name)
        if LH < 17:
            W = 190
            Lng = 16
        if LH > 17:
            W = 130
            Lng = 24

        return W, Lng, fp,
    def RemoveOccasion(name):
        name = name.replace("<HBD>", "")
        name = name.replace("<HAA>", "")
        name = name.replace("<IRO>", "")
        return name

    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        df = pandas.read_excel(excel_file)

        for i in range(len(df)):
            name = (df.loc[i, df.columns[0]])
            no = (df.loc[i, df.columns[1]])
            x = 0
            name = OShortName(name)

            while x < no:

                CH = 77

                x = x + 1
                y = y + 1
                qrimg = MakeQR(name)

                if name.find("<HBD>") != -1:
                    fp = open(hbd_path, "rb")
                    CH = 177
                    LH = len(name)
                    Ltype = 1
                    if LH < 17:
                        W = 190
                        Lng = 16
                    if LH > 17:
                        W = 130
                        Lng = 24
                    if LH > 57:
                        W = 100
                        Lng = 31
                elif name.find("<HAA>") != -1:
                    fp = open(anni_path, "rb")
                    CH = 177
                    LH = len(name)
                    Ltype = 1
                    if LH < 17:
                        W = 190
                        Lng = 16
                    if LH > 17:
                        W = 130
                        Lng = 24
                    if LH > 57:
                        W = 100
                        Lng = 31
                elif name.find("<IRO>") != -1:
                    fp = open(iro_path, "rb")
                    CH = 177
                    LH = len(name)
                    Ltype = 1
                    if LH < 17:
                        W = 190
                        Lng = 16
                    if LH > 17:
                        W = 130
                        Lng = 23
                    if LH > 57:
                        W = 100
                        Lng = 31
                else:
                    fp = open(wlf_path, "rb")
                    LH = len(name)
                    CH = 77
                    Ltype = 0
                    if LH < 33:
                        W = 190
                        Lng = 16
                    if LH > 32:
                        W = 130
                        Lng = 24
                    if LH > 72:
                        W = 100
                        Lng = 31

                wrname = RemoveOccasion(name)
                name = name.upper()
                # print(wrname)

                if name.isascii() is False:

                    para = textwrap.wrap(wrname, width=Lng)
                    MAX_W, MAX_H = 2000, 1000
                    wim = PIL.Image.open(white_path)
                    draw = ImageDraw.Draw(wim)
                    current_h = 0
                    current_h, pad = CH, -5

                    Lng = 1000
                    font = ImageFont.truetype(arial_path, W)

                    for line in para:
                        w, h = draw.textsize(line, font=font)
                        draw.text(((MAX_W - w) / 2, current_h), line, font=font, fill=(0, 0, 0))
                        current_h += h + pad
                    wim = trim(wim)

                    # fp = open("wlf.png", "rb")
                    im = PIL.Image.open(fp)

                    if Ltype == 0:
                        wim = resize_with_padding(wim, (1207, 329))
                        im.paste(wim, (18, 100))
                    if Ltype == 1:
                        wim = resize_with_padding(wim, (1207, 246))
                        im.paste(wim, (18, 184))

                    # im.show()

                else:

                    para = textwrap.wrap(wrname, width=Lng)
                    MAX_W, MAX_H = 2000, 1000
                    wim = PIL.Image.open(white_path)
                    draw = ImageDraw.Draw(wim)
                    current_h = 0
                    current_h, pad = CH, -5

                    Lng = 1000
                    font = ImageFont.truetype(bebas_path, W)

                    for line in para:
                        w, h = draw.textsize(line, font=font)
                        draw.text(((MAX_W - w) / 2, current_h), line, font=font, fill=(0, 0, 0))
                        current_h += h + pad
                    wim = trim(wim)
                    # wim = resize_with_padding(wim, (1207, 329))

                    # fp = open("wlf.png", "rb")
                    im = PIL.Image.open(fp)

                    if Ltype == 0:
                        wim = resize_with_padding(wim, (1207, 329))
                        im.paste(wim, (18, 100))
                    if Ltype == 1:
                        wim = resize_with_padding(wim, (1207, 246))
                        im.paste(wim, (18, 184))

                # im.paste(qrimg,(0,438))
                im.paste(qrimg, (25, 465))
                print(wrname)
                im.save(f'{temp_dir}/{y}.pdf', optimize=True, quality=50)
        #save_dir = "savepdf"

        mergedObject = PdfFileMerger()
        import os

        for fileNumber in (range(1, y + 1)):
            mergedObject.append(PdfFileReader(f'{temp_dir}/{fileNumber}.pdf'))
            print(y)

        from datetime import date
        pdate = date.today()
        mergedObject.write(f'pdfmergerout.pdf')

        import shutil
        shutil.rmtree(temp_dir)



        ######################
        """wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)"""

        
        ###########


        #return render(request, 'index.html')
        from django.shortcuts import redirect
        return redirect('/downloadpdf/')



